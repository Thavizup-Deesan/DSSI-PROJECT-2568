import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
from django.http import HttpResponse

def generate_po_excel(order):
    """
    Generate Excel file for Purchase Order
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"PO-{order.order_no}"

    # Styles
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = 'ใบสั่งซื้อ (Purchase Order)'
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = center_align

    # Info
    ws['A3'] = f"เลขที่เอกสาร: {order.order_no}"
    ws['D3'] = f"วันที่: {order.created_at.strftime('%d/%m/%Y')}"
    ws['A4'] = f"โครงการ: {order.project.project_name}"
    ws['A5'] = f"ผู้ขอซื้อ: {order.requester.full_name}"
    ws['A6'] = f"กรรมการตรวจรับ: {order.inspection_committee_name or '-'}"

    # Table Header
    headers = ['ลำดับ', 'รายการ', 'จำนวน', 'หน่วย', 'ราคา/หน่วย', 'ราคารวม']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col_num)
        cell.value = header
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = border

    # Items
    row_num = 9
    for idx, item in enumerate(order.items.all(), 1):
        ws.cell(row=row_num, column=1, value=idx).border = border
        ws.cell(row=row_num, column=2, value=item.material_name).border = border
        ws.cell(row=row_num, column=3, value=item.quantity).border = border
        ws.cell(row=row_num, column=4, value=item.unit).border = border
        ws.cell(row=row_num, column=5, value=float(item.unit_price)).border = border
        ws.cell(row=row_num, column=6, value=float(item.total_price)).border = border
        row_num += 1

    # Total
    ws.merge_cells(f'A{row_num}:E{row_num}')
    ws.cell(row=row_num, column=1, value='รวมทั้งสิ้น').alignment = Alignment(horizontal='right')
    ws.cell(row=row_num, column=1).font = bold_font
    ws.cell(row=row_num, column=6, value=float(order.total_amount)).font = bold_font
    ws.cell(row=row_num, column=6).border = border

    # Output
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
